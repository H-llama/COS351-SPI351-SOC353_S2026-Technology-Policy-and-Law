import os
import io
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scipy.stats import chi2_contingency, f_oneway
from sklearn.linear_model import LogisticRegression
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

os.makedirs("graphs", exist_ok=True)

def make_chart(fig, dpi=150):
    """Convert a matplotlib figure to an openpyxl Image (in-memory, no temp file)."""
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight")
    buf.seek(0)
    plt.close(fig)
    img = XLImage(buf)
    return img

# ============================================================
# LOAD + CLEAN DATA
# ============================================================

dataraw = pd.read_csv("scam_ads_labeled.csv")
data = dataraw.copy()
data["tier"] = data["tier"].astype(str).str.strip().str.lower()
data["brand"] = data["brand"].astype(str).str.strip().str.lower()
data["is_scam"] = data["is_scam"].astype(str).str.strip().str.lower() == "true"
data["brand_mentioned"] = data["brand_mentioned"].astype(str).str.strip().str.lower() == "true"
data["publisher_platform"] = data["publisher_platform"].astype(str).str.strip().str.lower()
data["start_date"] = pd.to_datetime(data["start_date"], errors="coerce")
data["month"] = data["start_date"].dt.month

# ============================================================
# COMPUTE ALL RESULTS
# ============================================================

# 1. Brand scam rates per tier
TIERS = ["fast_fashion", "luxury", "mid_tier"]
TIER_LABELS = {"fast_fashion": "Fast Fashion", "luxury": "Luxury", "mid_tier": "Mid Tier"}
brand_rates_by_tier = {}
for t in TIERS:
    subset = data[(data["tier"] == t) & (data["brand_mentioned"])]
    if len(subset) == 0:
        brand_rates_by_tier[t] = pd.DataFrame(columns=["Brand", "Scam Rate", "Scam Rate %"])
        continue
    br = subset.groupby("brand")["is_scam"].mean().sort_values(ascending=False).reset_index()
    br.columns = ["Brand", "Scam Rate"]
    br["Scam Rate %"] = (br["Scam Rate"] * 100).round(1)
    brand_rates_by_tier[t] = br

# 2. Chi-square test
contingency = pd.crosstab(data["tier"], data["is_scam"])
chi2, p, dof, expected = chi2_contingency(contingency)
n = contingency.values.sum()
k = min(contingency.shape)
cramers_v = np.sqrt(chi2 / (n * (k - 1)))

chi2_results = pd.DataFrame({
    "Metric": ["Chi2 Statistic", "p-value", "Degrees of Freedom", "Cramér's V", "N (total obs)"],
    "Value": [round(chi2, 4), round(p, 6), dof, round(cramers_v, 4), int(n)]
})

# 3. ANOVA
lux = data[data["tier"] == "luxury"]["criteria_met"].dropna()
mid = data[data["tier"] == "mid_tier"]["criteria_met"].dropna()
ff_crit = data[data["tier"] == "fast_fashion"]["criteria_met"].dropna()
f_stat, p_val = f_oneway(lux, mid, ff_crit)

anova_results = pd.DataFrame({
    "Metric": ["F-Statistic", "p-value", "n (luxury)", "n (mid_tier)", "n (fast_fashion)"],
    "Value": [round(f_stat, 4), round(p_val, 6), len(lux), len(mid), len(ff_crit)]
})

# 4. Logistic Regression
df_lr = dataraw.copy()
df_lr["is_scam"] = df_lr["is_scam"].astype(str).str.lower() == "true"
df_lr = pd.get_dummies(df_lr, columns=["tier"], drop_first=True)
features = [
    "c1_domain_mismatch", "c2_extreme_discount", "c3_urgency_language",
    "c4_category_mismatch", "c5_low_likes", "holiday_window"
] + [c for c in df_lr.columns if "tier_" in c]
df_lr[features] = df_lr[features].fillna(0)
X = df_lr[features]
y = df_lr["is_scam"].astype(int)
model = LogisticRegression(max_iter=1000)
model.fit(X, y)
coef_series = pd.Series(model.coef_[0], index=X.columns).sort_values(ascending=False)
logit_results = coef_series.reset_index()
logit_results.columns = ["Feature", "Coefficient"]
logit_results["Coefficient"] = logit_results["Coefficient"].round(4)

# 5. Platform distribution
platform_dist = pd.crosstab(
    data["publisher_platform"], data["tier"],
    values=data["is_scam"], aggfunc="mean"
).round(4).reset_index()

# 6. Temporal analysis
holiday = data[data["month"].isin([11, 12])]
non_holiday = data[~data["month"].isin([11, 12])]
holiday_rate = holiday.groupby("tier")["is_scam"].mean().round(4)
non_holiday_rate = non_holiday.groupby("tier")["is_scam"].mean().round(4)
temporal_df = pd.DataFrame({
    "Tier": holiday_rate.index,
    "Holiday Scam Rate": holiday_rate.values,
    "Non-Holiday Scam Rate": non_holiday_rate.reindex(holiday_rate.index).values,
})
temporal_df["Difference"] = (temporal_df["Holiday Scam Rate"] - temporal_df["Non-Holiday Scam Rate"]).round(4)

# 7. Zara coordinated campaign
zara = data[data["brand"].str.contains("zara", na=False)]
dup_ads = zara[zara.duplicated(subset=["ad_text"], keep=False)][["page_name", "ad_text"]].copy()

# 8. Summary
summary_df = pd.DataFrame({
    "Metric": ["Total Rows", "Total Scams", "Overall Scam Rate",
               "Fast Fashion Rows", "Luxury Rows", "Mid Tier Rows"],
    "Value": [
        len(data),
        int(data["is_scam"].sum()),
        f"{data['is_scam'].mean()*100:.1f}%",
        len(data[data["tier"] == "fast_fashion"]),
        len(data[data["tier"] == "luxury"]),
        len(data[data["tier"] == "mid_tier"]),
    ]
})

# Scam rate by tier
tier_scam = data.groupby("tier")["is_scam"].mean().round(4).reset_index()
tier_scam.columns = ["Tier", "Scam Rate"]
tier_scam["Scam Rate %"] = (tier_scam["Scam Rate"] * 100).round(1)

# Criteria frequency — aggregate + per tier
criteria_cols = ["c1_domain_mismatch", "c2_extreme_discount", "c3_urgency_language",
                 "c4_category_mismatch", "c5_low_likes"]
criteria_freq = data[criteria_cols].sum().reset_index()
criteria_freq.columns = ["Criterion", "Count"]

criteria_freq_by_tier = {}
for t in TIERS:
    cf = data[data["tier"] == t][criteria_cols].sum().reset_index()
    cf.columns = ["Criterion", "Count"]
    criteria_freq_by_tier[t] = cf

# Monthly scam rate — aggregate + per tier
monthly_rate = data.groupby("month")["is_scam"].mean().round(4).reset_index()
monthly_rate.columns = ["Month", "Scam Rate"]

monthly_rate_by_tier = {}
for t in TIERS:
    mr = data[data["tier"] == t].groupby("month")["is_scam"].mean().round(4).reset_index()
    mr.columns = ["Month", "Scam Rate"]
    monthly_rate_by_tier[t] = mr

# Contingency table (for chi-square sheet)
contingency_reset = contingency.reset_index()

# Chi-square per tier (each tier vs the other two combined)
chi2_by_tier = {}
for t in TIERS:
    d_binary = data.copy()
    d_binary["is_target"] = d_binary["tier"] == t
    ct = pd.crosstab(d_binary["is_target"], d_binary["is_scam"])
    if ct.shape == (2, 2):
        c2, pv, df2, _ = chi2_contingency(ct)
        n2 = ct.values.sum()
        cv = np.sqrt(c2 / (n2 * 1))
        chi2_by_tier[t] = pd.DataFrame({
            "Metric": ["Chi2 Statistic", "p-value", "Degrees of Freedom", "Cramér's V", "N"],
            "Value": [round(c2, 4), round(pv, 6), df2, round(cv, 4), int(n2)]
        })
    else:
        chi2_by_tier[t] = pd.DataFrame({"Metric": ["Insufficient data"], "Value": ["-"]})

# Logistic regression per tier
logit_by_tier = {}
base_features = ["c1_domain_mismatch", "c2_extreme_discount", "c3_urgency_language",
                 "c4_category_mismatch", "c5_low_likes", "holiday_window"]
for t in TIERS:
    subset = dataraw[dataraw["tier"].astype(str).str.strip().str.lower() == t].copy()
    if len(subset) < 20:
        logit_by_tier[t] = pd.DataFrame({"Feature": ["Not enough data"], "Coefficient": ["-"]})
        continue
    subset["is_scam"] = subset["is_scam"].astype(str).str.lower() == "true"
    avail = [f for f in base_features if f in subset.columns]
    subset[avail] = subset[avail].fillna(0)
    Xt = subset[avail]
    yt = subset["is_scam"].astype(int)
    if yt.nunique() < 2:
        logit_by_tier[t] = pd.DataFrame({"Feature": ["Only one class in target"], "Coefficient": ["-"]})
        continue
    mt = LogisticRegression(max_iter=1000)
    mt.fit(Xt, yt)
    cs = pd.Series(mt.coef_[0], index=Xt.columns).sort_values(ascending=False).reset_index()
    cs.columns = ["Feature", "Coefficient"]
    cs["Coefficient"] = cs["Coefficient"].round(4)
    logit_by_tier[t] = cs

# ============================================================
# BUILD EXCEL WORKBOOK
# ============================================================

wb = Workbook()

HEADER_FILL = PatternFill("solid", start_color="1F3864")
SUBHEADER_FILL = PatternFill("solid", start_color="2E75B6")
ALT_FILL = PatternFill("solid", start_color="D6E4F0")
WHITE_FILL = PatternFill("solid", start_color="FFFFFF")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="B0C4DE"),
    right=Side(style="thin", color="B0C4DE"),
    top=Side(style="thin", color="B0C4DE"),
    bottom=Side(style="thin", color="B0C4DE"),
)

def style_header(cell, sub=False):
    cell.fill = SUBHEADER_FILL if sub else HEADER_FILL
    cell.font = SUBHEADER_FONT if sub else HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER

def style_cell(cell, bold=False, alt=False):
    cell.fill = ALT_FILL if alt else WHITE_FILL
    cell.font = BOLD_FONT if bold else BODY_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = THIN_BORDER

def write_table(ws, df, start_row, title=None, title_col=1):
    if title:
        title_cell = ws.cell(row=start_row, column=title_col, value=title)
        title_cell.font = Font(name="Arial", bold=True, size=12, color="1F3864")
        title_cell.alignment = Alignment(horizontal="left")
        start_row += 1

    # Headers
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=col_name)
        style_header(cell, sub=True)

    # Data rows
    for row_idx, row in enumerate(df.itertuples(index=False), 1):
        alt = row_idx % 2 == 0
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=start_row + row_idx, column=col_idx, value=value)
            style_cell(cell, alt=alt)

    return start_row + len(df) + 2  # next free row

def autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

# ---- Sheet 1: Summary ----
ws1 = wb.active
ws1.title = "Summary"
ws1.freeze_panes = "A2"

# Big title
ws1.merge_cells("A1:D1")
title_cell = ws1["A1"]
title_cell.value = "Scam Ads Analysis — Summary Report"
title_cell.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
title_cell.fill = HEADER_FILL
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 30

row = 3
row = write_table(ws1, summary_df, row, title="📊 Overall Summary")
row = write_table(ws1, tier_scam, row, title="📊 Scam Rate by Tier")

row = write_table(ws1, criteria_freq, row, title="📊 Criteria Frequency — All Tiers (Aggregate)")
for t in TIERS:
    row = write_table(ws1, criteria_freq_by_tier[t], row, title=f"📊 Criteria Frequency — {TIER_LABELS[t]}")

row = write_table(ws1, monthly_rate, row, title="📊 Monthly Scam Rate — All Tiers (Aggregate)")
for t in TIERS:
    row = write_table(ws1, monthly_rate_by_tier[t], row, title=f"📊 Monthly Scam Rate — {TIER_LABELS[t]}")
autofit(ws1)

# ---- Sheet 2: Brand Scam Rates (all tiers) ----
ws2 = wb.create_sheet("Brand Scam Rates")
ws2.merge_cells("A1:D1")
c = ws2["A1"]
c.value = "Brand Scam Rates by Tier (brand_mentioned=True)"
c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
c.fill = HEADER_FILL
c.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

row2 = 3
for t in TIERS:
    label = TIER_LABELS[t]
    br = brand_rates_by_tier[t]
    row2 = write_table(ws2, br, row2, title=f"📊 {label} — Brand Scam Rates")
autofit(ws2)

# ---- Sheet 3: Statistical Tests ----
ws3 = wb.create_sheet("Statistical Tests")
ws3.merge_cells("A1:D1")
c = ws3["A1"]
c.value = "Statistical Test Results"
c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
c.fill = HEADER_FILL
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

row = 3
row = write_table(ws3, chi2_results, row, title="Chi-Square Test — All Tiers (Aggregate, Tier vs Scam)")
for t in TIERS:
    row = write_table(ws3, chi2_by_tier[t], row, title=f"Chi-Square Test — {TIER_LABELS[t]} vs Others")
row = write_table(ws3, anova_results, row, title="ANOVA (Criteria Met Across Tiers)")
row = write_table(ws3, contingency_reset, row, title="Contingency Table (Tier × Is Scam)")
autofit(ws3)

ws4 = wb.create_sheet("Logistic Regression")
ws4.merge_cells("A1:C1")
c = ws4["A1"]
c.value = "Logistic Regression — Feature Coefficients (predicting is_scam)"
c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
c.fill = HEADER_FILL
c.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 28
row4 = 3
row4 = write_table(ws4, logit_results, row4, title="All Tiers — Aggregate Model")
for t in TIERS:
    row4 = write_table(ws4, logit_by_tier[t], row4, title=f"{TIER_LABELS[t]} — Tier-Specific Model")
autofit(ws4)

# ---- Sheet 5: Platform Distribution ----
ws5 = wb.create_sheet("Platform Distribution")
ws5.merge_cells("A1:F1")
c = ws5["A1"]
c.value = "Scam Rate by Platform & Tier"
c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
c.fill = HEADER_FILL
c.alignment = Alignment(horizontal="center", vertical="center")
ws5.row_dimensions[1].height = 28
write_table(ws5, platform_dist, 3)
autofit(ws5)

# ---- Sheet 6: Temporal Analysis ----
ws6 = wb.create_sheet("Temporal Analysis")
ws6.merge_cells("A1:E1")
c = ws6["A1"]
c.value = "Temporal Clustering — Holiday vs Non-Holiday Scam Rates"
c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
c.fill = HEADER_FILL
c.alignment = Alignment(horizontal="center", vertical="center")
ws6.row_dimensions[1].height = 28
write_table(ws6, temporal_df, 3)
autofit(ws6)

# ---- Sheet 7: Zara Duplicated Ads ----
ws7 = wb.create_sheet("Zara Duplicated Ads")
ws7.merge_cells("A1:C1")
c = ws7["A1"]
c.value = "Zara — Suspected Coordinated Campaign (Duplicated Ad Text)"
c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
c.fill = HEADER_FILL
c.alignment = Alignment(horizontal="center", vertical="center")
ws7.row_dimensions[1].height = 28

if len(dup_ads) > 0:
    write_table(ws7, dup_ads.reset_index(drop=True), 3)
else:
    ws7["A3"] = "No duplicated ads found."
    ws7["A3"].font = BODY_FONT
autofit(ws7)
# Wrap text for ad_text column
for row in ws7.iter_rows(min_row=4):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
ws7.row_dimensions  # let rows auto-size via wrap

# ============================================================
# 8. GRAPHS SHEET
# ============================================================

CHART_COLOR = "#2E75B6"
ACCENT_COLORS = ["#1F3864", "#2E75B6", "#70AD47", "#ED7D31", "#FFC000"]

# --- Chart 1: Scam Rate by Tier ---
tier_plot = data.groupby("tier")["is_scam"].mean().sort_values(ascending=False)
fig1, ax1 = plt.subplots(figsize=(7, 4))
bars = ax1.bar(tier_plot.index, tier_plot.values * 100, color=ACCENT_COLORS[:len(tier_plot)])
ax1.set_title("Scam Rate by Tier", fontsize=13, fontweight="bold", pad=12)
ax1.set_ylabel("Scam Rate (%)")
ax1.set_xlabel("Tier")
for bar in bars:
    ax1.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.5,
             f"{bar.get_height():.1f}%", ha="center", va="bottom", fontsize=9)
ax1.spines[["top", "right"]].set_visible(False)
ax1.set_ylim(0, max(tier_plot.values * 100) * 1.2)
fig1.tight_layout()
img1 = make_chart(fig1)

# --- Chart 2: Criteria Frequency ---
crit_data = data[criteria_cols].sum().sort_values(ascending=False)
fig2, ax2 = plt.subplots(figsize=(7, 4))
bars2 = ax2.barh(crit_data.index[::-1], crit_data.values[::-1], color=CHART_COLOR)
ax2.set_title("Criteria Frequency", fontsize=13, fontweight="bold", pad=12)
ax2.set_xlabel("Count")
for bar in bars2:
    ax2.text(bar.get_width() + 0.5, bar.get_y() + bar.get_height() / 2,
             str(int(bar.get_width())), va="center", fontsize=9)
ax2.spines[["top", "right"]].set_visible(False)
fig2.tight_layout()
img2 = make_chart(fig2)

# --- Chart 3: Scam Rate Over Time ---
monthly_plot = data.groupby("month")["is_scam"].mean()
fig3, ax3 = plt.subplots(figsize=(7, 4))
ax3.plot(monthly_plot.index, monthly_plot.values * 100, marker="o",
         color=CHART_COLOR, linewidth=2, markersize=6)
ax3.fill_between(monthly_plot.index, monthly_plot.values * 100, alpha=0.15, color=CHART_COLOR)
ax3.set_title("Scam Rate Over Time", fontsize=13, fontweight="bold", pad=12)
ax3.set_xlabel("Month")
ax3.set_ylabel("Scam Rate (%)")
ax3.set_xticks(monthly_plot.index)
ax3.spines[["top", "right"]].set_visible(False)
fig3.tight_layout()
img3 = make_chart(fig3)

# --- Chart 4: Holiday vs Non-Holiday Scam Rate ---
x = np.arange(len(temporal_df))
width = 0.35
fig4, ax4 = plt.subplots(figsize=(7, 4))
b1 = ax4.bar(x - width/2, temporal_df["Holiday Scam Rate"] * 100, width,
             label="Holiday (Nov–Dec)", color="#ED7D31")
b2 = ax4.bar(x + width/2, temporal_df["Non-Holiday Scam Rate"] * 100, width,
             label="Non-Holiday", color=CHART_COLOR)
ax4.set_title("Holiday vs Non-Holiday Scam Rate by Tier", fontsize=13, fontweight="bold", pad=12)
ax4.set_ylabel("Scam Rate (%)")
ax4.set_xticks(x)
ax4.set_xticklabels(temporal_df["Tier"])
ax4.legend()
ax4.spines[["top", "right"]].set_visible(False)
fig4.tight_layout()
img4 = make_chart(fig4)

# --- Chart 5: Logistic Regression Coefficients ---
fig5, ax5 = plt.subplots(figsize=(7, 5))
colors = ["#C00000" if v > 0 else "#2E75B6" for v in logit_results["Coefficient"]]
ax5.barh(logit_results["Feature"][::-1], logit_results["Coefficient"][::-1], color=colors[::-1])
ax5.axvline(0, color="black", linewidth=0.8)
ax5.set_title("Logistic Regression Coefficients\n(Red = increases scam probability)",
              fontsize=12, fontweight="bold", pad=12)
ax5.set_xlabel("Coefficient")
ax5.spines[["top", "right"]].set_visible(False)
fig5.tight_layout()
img5 = make_chart(fig5)

# --- Chart 6: Brand Scam Rates per Tier (top 5 brands each) ---
TIER_COLORS = {"fast_fashion": "#2E75B6", "luxury": "#1F3864", "mid_tier": "#70AD47"}
fig6, axes = plt.subplots(1, 3, figsize=(14, 5), sharey=False)
for ax, t in zip(axes, TIERS):
    br = brand_rates_by_tier[t].head(5)
    label = TIER_LABELS[t]
    if len(br) == 0:
        ax.text(0.5, 0.5, "No data", ha="center", va="center", transform=ax.transAxes)
        ax.set_title(label, fontsize=11, fontweight="bold")
        continue
    ax.barh(br["Brand"][::-1], br["Scam Rate %"][::-1], color=TIER_COLORS[t])
    ax.set_title(f"{label}\n(Top {len(br)} brands)", fontsize=11, fontweight="bold")
    ax.set_xlabel("Scam Rate (%)")
    ax.spines[["top", "right"]].set_visible(False)
fig6.suptitle("Brand Scam Rates by Tier", fontsize=13, fontweight="bold", y=1.02)
fig6.tight_layout()
img6 = make_chart(fig6)

# --- Build the Graphs sheet ---
ws_graphs = wb.create_sheet("Graphs")
ws_graphs.merge_cells("A1:N1")
c = ws_graphs["A1"]
c.value = "Visual Analysis — All Charts"
c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
c.fill = HEADER_FILL
c.alignment = Alignment(horizontal="center", vertical="center")
ws_graphs.row_dimensions[1].height = 28

CHART_W, CHART_H = 480, 280

def place_image(ws, img, col_letter, row):
    img.width = CHART_W
    img.height = CHART_H
    ws.add_image(img, f"{col_letter}{row}")

CHART_W_WIDE = 960

def place_image_wide(ws, img, col_letter, row):
    img.width = CHART_W_WIDE
    img.height = CHART_H
    ws.add_image(img, f"{col_letter}{row}")

place_image(ws_graphs, img1, "A", 3)
place_image(ws_graphs, img2, "I", 3)
place_image(ws_graphs, img3, "A", 22)
place_image(ws_graphs, img4, "I", 22)
place_image(ws_graphs, img5, "A", 41)
place_image_wide(ws_graphs, img6, "I", 41)

# ============================================================
# SAVE
# ============================================================

out_path = "scam_ads_analysis.xlsx"
wb.save(out_path)
print(f"Saved to {out_path}")